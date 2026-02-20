#app/excel/py_xlsx.py 

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path
from datetime import datetime
from sqlalchemy.orm import Session
from app.db.model import Data, Product, Characteristic, Bank


RUSSIAN_CHAR_NAMES = {
    "type": "–¢–∏–ø –∫–∞—Ä—Ç—ã",
    "currency": "–í–∞–ª—é—Ç–∞", 
    "validity": "–°—Ä–æ–∫ –¥–µ–π—Å—Ç–≤–∏—è",
    "maintenance_cost": "–û–±—Å–ª—É–∂–∏–≤–∞–Ω–∏–µ",
    "free_conditions": "–ë–µ—Å–ø–ª–∞—Ç–Ω–æ –ø—Ä–∏",
    "sms_notification": "–°–ú–° —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è",
    "atm_limit_own": "–õ–∏–º–∏—Ç ATM —Å–≤–æ–µ–≥–æ",
    "atm_limit_other": "–õ–∏–º–∏—Ç ATM –¥—Ä—É–≥–∏—Ö",
    "loyalty_program": "–ü—Ä–æ–≥—Ä–∞–º–º–∞ –ª–æ—è–ª—å–Ω–æ—Å—Ç–∏",
    "interest_rate": "% –Ω–∞ –æ—Å—Ç–∞—Ç–æ–∫",
    "additional": "–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ",
}


def _get_russian_char_name(char_name: str) -> str:
    if char_name in RUSSIAN_CHAR_NAMES:
        return RUSSIAN_CHAR_NAMES[char_name]
    return char_name


def create_bank_excel_report(
    db: Session,
    user_id: int,
    product_ids: list[int],
    char_ids: list[int],
    output_dir: str = "./reports/"
) -> str:
    
    try:
        # –ü–æ–ª—É—á–∞–µ–º –¥–∞–Ω–Ω—ã–µ
        data_records = db.query(Data).filter(
            Data.user_id == user_id,
            Data.product_id.in_(product_ids)
        ).all()
        
        print(f"–ù–∞–π–¥–µ–Ω–æ {len(data_records)} –∑–∞–ø–∏—Å–µ–π –≤ –ë–î")
        
        if not data_records:
            print("–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –æ—Ç—á–µ—Ç–∞")
            return None
        
        # –ü–æ–ª—É—á–∞–µ–º –º–µ—Ç–∞–¥–∞–Ω–Ω—ã–µ
        products = db.query(Product).filter(Product.id.in_(product_ids)).all()
        chars = db.query(Characteristic).filter(Characteristic.id.in_(char_ids)).all()
        banks = db.query(Bank).all()
        
        print(f"–•–∞—Ä–∞–∫—Ç–µ—Ä–∏—Å—Ç–∏–∫: {len(chars)}, –ü—Ä–æ–¥—É–∫—Ç–æ–≤: {len(products)}")
        
        product_map = {p.id: p for p in products}
        char_map = {c.id: c for c in chars}
        bank_map = {b.id: b for b in banks}
        
        # –°—Ç—Ä—É–∫—Ç—É—Ä–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ
        structured_data = {}
        for record in data_records:
            if record.product_id not in structured_data:
                structured_data[record.product_id] = {}
            structured_data[record.product_id][record.characteristic_id] = record.value
        
        table_data = []
        product_urls = {}  # –î–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è URL –ø—Ä–æ–¥—É–∫—Ç–æ–≤
        
        for char in chars:
            char_display_name = _get_russian_char_name(char.name)
            
            row = {
                "–•–∞—Ä–∞–∫—Ç–µ—Ä–∏—Å—Ç–∏–∫–∞": char_display_name
            }
            
            for product in products:
                value = structured_data.get(product.id, {}).get(char.id, "‚Äî")
                
                bank_name = bank_map.get(product.bank_id, None)
                bank_name = bank_name.name if bank_name else "Unknown"
                col_name = f"{bank_name}\n{product.name}"
                
                row[col_name] = value
                
                if product.id not in product_urls:
                    product_urls[product.id] = {
                        "url": product.url,
                        "name": product.name,
                        "bank": bank_name,
                        "col_name": col_name
                    }
            
            table_data.append(row)
        
        # –°–æ–∑–¥–∞—ë–º DataFrame
        df = pd.DataFrame(table_data)
        
        # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –∏–º—è —Ñ–∞–π–ª–∞
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"–ü–∞—Ä—Å–∏–Ω–≥_–ö–∞—Ä—Ç_{timestamp}.xlsx"
        
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        filepath = output_path / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(
                writer,
                sheet_name='–°—Ä–∞–≤–Ω–µ–Ω–∏–µ',
                index=False,
                startrow=2,
                header=False  
            )
            
            worksheet = writer.sheets['–°—Ä–∞–≤–Ω–µ–Ω–∏–µ']
            

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                cell.value = df.columns[col - 1]
            
            cell = worksheet.cell(row=2, column=1)
            cell.value = "–°—Å—ã–ª–∫–∞ –Ω–∞ —Å–∞–π—Ç"
            cell.font = Font(bold=True, size=10, color="0563C1")
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.fill = PatternFill(start_color="E7F0FF", end_color="E7F0FF", fill_type="solid")
            
            # –î–æ–±–∞–≤–ª—è–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –¥–ª—è –∫–∞–∂–¥–æ–≥–æ –ø—Ä–æ–¥—É–∫—Ç–∞ (—Å—Ç—Ä–æ–∫–∞ 2)
            for col_idx, product in enumerate(products, 1):
                if product.id in product_urls:
                    product_info = product_urls[product.id]
                    url = product_info["url"]
                    name = product_info["name"]
                    bank = product_info["bank"]
                    
                    cell = worksheet.cell(row=2, column=col_idx + 1)
                    
                    if url:
                        cell.hyperlink = url
                        cell.value = f"üîó {bank}\n{name}"
                        cell.font = Font(color="0563C1", underline="single", size=10, bold=True)
                    else:
                        cell.value = f"{bank}\n{name}"
                        cell.font = Font(bold=True, size=10)
                    
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    cell.fill = PatternFill(start_color="E7F0FF", end_color="E7F0FF", fill_type="solid")
            

            data_start_row = 3  
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=data_start_row), 1):
                for col_idx, cell in enumerate(row):
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    
                    # –ß–µ—Ä–µ–¥—É–µ–º —Ü–≤–µ—Ç–∞ —Å—Ç—Ä–æ–∫
                    if (row_idx + 1) % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    # –í—ã–¥–µ–ª—è–µ–º –ø—É—Å—Ç—ã–µ —è—á–µ–π–∫–∏
                    if cell.value is None or cell.value == "‚Äî":
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # –®–∏—Ä–∏–Ω–∞ –∫–æ–ª–æ–Ω–æ–∫
            worksheet.column_dimensions['A'].width = 30
            for col_idx in range(2, len(df.columns) + 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 25
            
            # –í—ã—Å–æ—Ç–∞ —Å—Ç—Ä–æ–∫
            worksheet.row_dimensions[1].height = 35  # –®–∞–ø–∫–∞ (—Å–∏–Ω—è—è)
            worksheet.row_dimensions[2].height = 40  # –°—Å—ã–ª–∫–∏ (—Å–≤–µ—Ç–ª–æ-–≥–æ–ª—É–±–∞—è)
            worksheet.row_dimensions[3].height = 20  # –ó–∞–≥–æ–ª–æ–≤–∫–∏ pandas
            for row in range(4, len(df) + 4): 
                worksheet.row_dimensions[row].height = 30
            

        
        print(f"Excel —Å–æ–∑–¥–∞–Ω: {filepath}")
        return str(filepath)
        
    except Exception as e:
        print(f"!!! –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ Excel: {e}")
        import traceback
        traceback.print_exc()
        return None